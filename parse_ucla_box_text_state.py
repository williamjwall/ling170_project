#!/usr/bin/env python3
"""
Scrape UCLA Box shared folder and parse TextGrid text/state into one dataset.

Corpus layout (UCLA Speaker Variability Database — see readme_data.txt):
  Filenames: {speaker}{session}_{task}[_{FAVE|darla}].TextGrid
  - speaker: numeric id (folder name matches this id)
  - session: A/B/C for the three main sessions, or D for an extra recording session
    (public_database_speaker_info.xlsx Notes sheet)
  - task: instructions, neutral, happy, phonecall, annoyed, video, sentences, vowels
  - optional suffix: _FAVE or _darla = forced-alignment output tiers; bare .TextGrid =
    orthographic tier used as alignment input (readme_data.txt)

This repo may only contain a Box *sample* (subset of speakers/files); the parser does
not assume the full 202-speaker tree.

Downloads: per speaker folder, TextGrids are fetched in parallel (--workers, default 8)
with retries. Folder names may zero-pad speaker ids (007); CSV speaker_id is unpadded
to match public_database_speaker_info.xlsx.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Tuple
from urllib.parse import quote
from urllib.request import Request, urlopen


SHARED_NAME = "iu8mmvuh16a19e8phjg0ol8gub0zyzf9"
ROOT_FOLDER_ID = "38657103233"
BASE_FOLDER_URL = f"https://ucla.app.box.com/s/{SHARED_NAME}/folder"
DOWNLOAD_URL = "https://ucla.app.box.com/index.php?rm=box_download_shared_file"
OUT_DIR = Path("ucla_box_parsed")
DEFAULT_SPEAKER_INFO = Path("public_database_speaker_info.xlsx")

# Filename: digits + session letter + underscore + task [ + _aligner ]
_UCLA_TEXTGRID_RE = re.compile(
    r"^(?P<speaker>\d+)(?P<session>[ABCDabcd])_(?P<rest>.+)$", re.I
)


def parse_numeric_speaker_id(label: str) -> Optional[int]:
    """If label is digits-only (e.g. 007), return int; else None."""
    s = label.strip()
    if s.isdigit():
        return int(s)
    return None


def speaker_labels_conflict(folder_speaker: str, filename_speaker: str) -> bool:
    """True only if both parse as integers and disagree (7 vs 007 is not a conflict)."""
    a = parse_numeric_speaker_id(folder_speaker)
    b = parse_numeric_speaker_id(filename_speaker)
    if a is not None and b is not None:
        return a != b
    if not folder_speaker or not filename_speaker:
        return False
    return folder_speaker.strip() != filename_speaker.strip()


def canonical_speaker_id(folder_speaker: str, filename_speaker: str) -> str:
    """Stable id: unpadded digits when folder is numeric (matches Excel speakerID)."""
    if parse_numeric_speaker_id(folder_speaker) is not None:
        return str(int(folder_speaker.strip()))
    if filename_speaker.strip().isdigit():
        return str(int(filename_speaker.strip()))
    return filename_speaker.strip() or folder_speaker.strip()


@dataclass
class PendingTextgrid:
    folder_id: str
    file_id: str
    file_name: str
    extension: str
    speaker_id: str
    session: str
    task: str
    variant: str
    textgrid_role: str
    session_key: str
    meta: Dict[str, Any]


@dataclass
class FileRow:
    speaker_id: str
    folder_id: str
    file_id: str
    file_name: str
    extension: str
    session: str
    task: str
    variant: str
    textgrid_role: str
    text: str
    # Composite key matching historical CSVs / quick filtering (speaker + session)
    session_key: str = field(default="")
    meta: Dict[str, Any] = field(default_factory=dict)


def fetch_text(url: str, timeout: int = 30) -> str:
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=timeout) as resp:
        return resp.read().decode("utf-8", errors="replace")


def extract_poststream_json(html: str) -> Dict:
    match = re.search(r"Box\.postStreamData\s*=\s*(\{.*?\});</script>", html, re.S)
    if not match:
        raise RuntimeError("Could not find Box.postStreamData payload")
    return json.loads(match.group(1))


KNOWN_TASKS = frozenset(
    {
        "instructions",
        "neutral",
        "happy",
        "phonecall",
        "annoyed",
        "video",
        "sentences",
        "vowels",
    }
)


def parse_ucla_textgrid_stem(
    file_name: str,
) -> Tuple[str, str, str, str, str]:
    """
    Parse `{speaker}{A-D}_{task}[_{FAVE|darla}]` before extension.

    Returns (speaker_id, session_letter, task, variant, textgrid_role) where
    textgrid_role is orthographic | aligned_fave | aligned_darla | unknown.
    """
    base = file_name.rsplit(".", 1)[0]
    m = _UCLA_TEXTGRID_RE.match(base)
    if not m:
        return ("", "", "unknown", "", "unknown")

    speaker_id = m.group("speaker")
    session = m.group("session").upper()
    rest = m.group("rest")
    lower = rest.lower()

    variant = ""
    task_body = rest
    if lower.endswith("_fave"):
        task_body = rest[:-5]
        variant = "FAVE"
    elif lower.endswith("_darla"):
        task_body = rest[:-6]
        variant = "darla"

    if variant == "FAVE":
        role = "aligned_fave"
    elif variant == "darla":
        role = "aligned_darla"
    else:
        role = "orthographic"

    task = task_body.lower() if task_body else "unknown"
    return (speaker_id, session, task, variant, role)


def task_public_info_columns(task: str, session: str) -> Optional[Tuple[str, str]]:
    """Map task + filename session to (Session_col, Clipping_col) in public_database_speaker_info."""
    t = task.lower()
    s = session.upper()
    if t == "instructions":
        return ("Instructions_Session", "Instructions_Clipping")
    if t == "neutral":
        return ("PromptsN_Session", "PromptsN_Clipping")
    if t == "happy":
        return ("PromptsH_Session", "PromptsH_Clipping")
    if t == "annoyed":
        return ("PromptsA_Session", "PromptsA_Clipping")
    if t == "phonecall":
        return ("Phonecall_Session", "Phonecall_Clipping")
    if t == "video":
        return ("Video_Session", "Video_Clipping")
    if t == "sentences":
        return (f"{s}_Sentences_Session", f"{s}_Sentences_Clipping")
    if t == "vowels":
        return (f"{s}_Vowels_Session", f"{s}_Vowels_Clipping")
    return None


def try_load_public_speaker_info(
    path: Path,
) -> Optional[Dict[int, Dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    if not path.is_file():
        return None
    wb = load_workbook(path, read_only=True, data_only=True)
    if "MetaData" not in wb.sheetnames:
        wb.close()
        return None
    ws = wb["MetaData"]
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        wb.close()
        return None
    idx: Dict[str, int] = {}
    for i, name in enumerate(header_row):
        if name is not None and str(name).strip():
            idx[str(name)] = i

    out: Dict[int, Dict[str, Any]] = {}
    for row in rows:
        if not row or row[0] is None:
            continue
        try:
            sid = int(row[0])
        except (TypeError, ValueError):
            continue
        rec: Dict[str, Any] = {}
        for key, i in idx.items():
            if i < len(row):
                v = row[i]
                if isinstance(v, str):
                    v = v.strip()
                rec[key] = v
        out[sid] = rec
    wb.close()
    return out


def merge_public_info_row(
    table: Mapping[int, Dict[str, Any]],
    speaker_id: str,
    task: str,
    session: str,
) -> Dict[str, Any]:
    try:
        sid = int(speaker_id)
    except ValueError:
        return {}
    row = table.get(sid)
    if not row:
        return {}

    def cell(k: str) -> Any:
        v = row.get(k)
        return v.strip() if isinstance(v, str) else v

    out: Dict[str, Any] = {
        "info_sex": cell("sex"),
        "info_age": row.get("age"),
        "info_l1_english": cell("L1=English"),
        "info_l1_other": cell("L1=other"),
        "info_l2_english_l1": cell("L2=EnglishL1"),
        "info_l2_english_aoa": cell("L2=EnglishAoA"),
    }
    pair = task_public_info_columns(task, session)
    if pair:
        scol, ccol = pair
        out["info_db_session"] = cell(scol)
        out["info_db_clipping"] = cell(ccol)
    return out


def extract_textgrid_text(content: str) -> str:
    # Keep text-bearing interval/point payloads only.
    lines = []
    for line in content.splitlines():
        line = line.strip()
        if line.startswith("text = "):
            m = re.match(r'text = "(.*)"$', line)
            if m:
                txt = m.group(1).strip()
                if txt:
                    lines.append(txt)
    return " ".join(lines)


def iter_folder_items(folder_id: str) -> Iterable[Dict]:
    page = 1
    while True:
        html = fetch_text(f"{BASE_FOLDER_URL}/{folder_id}?page={page}")
        payload = extract_poststream_json(html)
        info = payload["/app-api/enduserapp/shared-folder"]
        for item in info.get("items", []):
            yield item
        page_count = int(info.get("pageCount", 1))
        if page >= page_count:
            break
        page += 1
        time.sleep(0.1)


def download_shared_file(shared_name: str, file_id: str) -> str:
    url = (
        f"{DOWNLOAD_URL}&shared_name={quote(shared_name)}&file_id=f_{quote(file_id)}"
    )
    return fetch_text(url, timeout=60)


def download_parse_textgrid(
    shared_name: str,
    file_id: str,
    *,
    retries: int,
    backoff_sec: float,
) -> str:
    last_exc: Optional[BaseException] = None
    for attempt in range(max(1, retries)):
        try:
            content = download_shared_file(shared_name, file_id)
            return extract_textgrid_text(content)
        except BaseException as exc:
            last_exc = exc
            if attempt + 1 < retries:
                time.sleep(backoff_sec * (attempt + 1))
    return f"[ERROR: {last_exc}]"


def run_pending_downloads(
    shared_name: str,
    pending: List[PendingTextgrid],
    workers: int,
    retries: int,
    backoff_sec: float,
) -> List[str]:
    if not pending:
        return []
    if workers <= 1:
        return [
            download_parse_textgrid(
                shared_name, p.file_id, retries=retries, backoff_sec=backoff_sec
            )
            for p in pending
        ]

    def work(fid: str) -> str:
        return download_parse_textgrid(
            shared_name, fid, retries=retries, backoff_sec=backoff_sec
        )

    texts: List[Optional[str]] = [None] * len(pending)
    with ThreadPoolExecutor(max_workers=workers) as ex:
        future_map = {ex.submit(work, p.file_id): i for i, p in enumerate(pending)}
        for fut in as_completed(future_map):
            i = future_map[fut]
            texts[i] = fut.result()
    return [t if t is not None else "[ERROR: missing result]" for t in texts]


INFO_FIELDNAMES = [
    "info_sex",
    "info_age",
    "info_l1_english",
    "info_l1_other",
    "info_l2_english_l1",
    "info_l2_english_aoa",
    "info_db_session",
    "info_db_clipping",
]


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=(
            "Scrape the UCLA Box shared folder and write TextGrid-derived text to CSV. "
            "Filenames follow {speaker}{A-D}_{task}[_{FAVE|darla}].TextGrid — see readme_data.txt."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Optional: place public_database_speaker_info.xlsx in the working directory and pass "
            "--with-speaker-info to attach demographics and per-task session/clipping from the "
            "MetaData sheet. Use --max-speakers / --max-files when the Box link only exposes a sample."
        ),
    )
    p.add_argument(
        "--out-dir",
        type=Path,
        default=OUT_DIR,
        help=f"Output directory (default: {OUT_DIR})",
    )
    p.add_argument(
        "--root-folder-id",
        default=ROOT_FOLDER_ID,
        help="Box shared-folder root id (default: full corpus sample root)",
    )
    p.add_argument(
        "--max-speakers",
        type=int,
        default=None,
        metavar="N",
        help="Only process the first N speaker folders (useful for Box samples)",
    )
    p.add_argument(
        "--max-files",
        type=int,
        default=None,
        metavar="N",
        help="Stop after N TextGrid rows total",
    )
    p.add_argument(
        "--with-speaker-info",
        action="store_true",
        help="Merge columns from public_database_speaker_info.xlsx (MetaData sheet)",
    )
    p.add_argument(
        "--speaker-info",
        type=Path,
        default=DEFAULT_SPEAKER_INFO,
        help=f"Path to Excel speaker info (default: ./{DEFAULT_SPEAKER_INFO})",
    )
    p.add_argument(
        "--workers",
        type=int,
        default=8,
        metavar="N",
        help="Parallel TextGrid downloads per speaker folder (default: 8; use 1 for sequential)",
    )
    p.add_argument(
        "--retries",
        type=int,
        default=3,
        metavar="N",
        help="Download attempts per file on failure (default: 3)",
    )
    p.add_argument(
        "--retry-backoff",
        type=float,
        default=0.75,
        metavar="SEC",
        help="Base sleep between retries, scaled by attempt (default: 0.75)",
    )
    return p


def main() -> None:
    args = build_arg_parser().parse_args()
    out_dir: Path = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    public_table: Optional[Dict[int, Dict[str, Any]]] = None
    if args.with_speaker_info:
        public_table = try_load_public_speaker_info(args.speaker_info)
        if public_table is None:
            print(
                "Warning: could not load speaker info "
                f"({args.speaker_info}); install openpyxl or check path. "
                "Continuing without info_* columns.",
                flush=True,
            )

    root_items = list(iter_folder_items(args.root_folder_id))
    speaker_folders = [i for i in root_items if i.get("type") == "folder"]
    if args.max_speakers is not None:
        speaker_folders = speaker_folders[: args.max_speakers]
    misc_files = [i for i in root_items if i.get("type") == "file"]

    rows: List[FileRow] = []
    stopped_early = False
    total_files = 0
    workers = max(1, args.workers)
    retries = max(1, args.retries)
    backoff = max(0.0, args.retry_backoff)

    for idx, folder in enumerate(speaker_folders, start=1):
        folder_speaker = folder.get("name", "").strip()
        folder_id = str(folder.get("id", ""))
        print(
            f"[{idx}/{len(speaker_folders)}] speaker {folder_speaker} ({folder_id})",
            flush=True,
        )
        pending: List[PendingTextgrid] = []
        for item in iter_folder_items(folder_id):
            if item.get("type") != "file":
                continue
            extension = (item.get("extension") or "").lower()
            if extension != "textgrid":
                continue

            file_id = str(item.get("id", ""))
            file_name = item.get("name", "")
            spk, session, task, variant, role = parse_ucla_textgrid_stem(file_name)

            if task != "unknown" and task not in KNOWN_TASKS:
                print(
                    f"  note: unrecognized task {task!r} in {file_name!r}",
                    flush=True,
                )

            if spk and folder_speaker and speaker_labels_conflict(folder_speaker, spk):
                print(
                    f"  warning: filename speaker {spk!r} != folder {folder_speaker!r} "
                    f"({file_name})",
                    flush=True,
                )

            speaker_id = canonical_speaker_id(folder_speaker, spk)
            session_key = f"{speaker_id}{session}" if speaker_id and session else ""

            meta: Dict[str, Any] = {}
            if public_table:
                meta = merge_public_info_row(public_table, speaker_id, task, session)

            pending.append(
                PendingTextgrid(
                    folder_id=folder_id,
                    file_id=file_id,
                    file_name=file_name,
                    extension=extension,
                    speaker_id=speaker_id,
                    session=session,
                    task=task,
                    variant=variant,
                    textgrid_role=role,
                    session_key=session_key,
                    meta=meta,
                )
            )
            total_files += 1

            if args.max_files is not None and total_files >= args.max_files:
                stopped_early = True
                break

        texts = run_pending_downloads(
            SHARED_NAME, pending, workers, retries, backoff
        )
        for p, parsed_text in zip(pending, texts):
            rows.append(
                FileRow(
                    speaker_id=p.speaker_id,
                    folder_id=p.folder_id,
                    file_id=p.file_id,
                    file_name=p.file_name,
                    extension=p.extension,
                    session=p.session,
                    task=p.task,
                    variant=p.variant,
                    textgrid_role=p.textgrid_role,
                    text=parsed_text,
                    session_key=p.session_key,
                    meta=p.meta,
                )
            )

        if stopped_early:
            break

    base_fields = [
        "speaker_id",
        "folder_id",
        "file_id",
        "file_name",
        "extension",
        "session",
        "session_key",
        "task",
        "variant",
        "textgrid_role",
        "text",
    ]
    fieldnames = base_fields + (INFO_FIELDNAMES if public_table else [])

    csv_path = out_dir / "ucla_text_state_parsed.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            d = {k: getattr(row, k) for k in base_fields}
            if public_table:
                for k in INFO_FIELDNAMES:
                    d[k] = row.meta.get(k, "")
            writer.writerow(d)

    files_csv = out_dir / "root_misc_files.csv"
    with files_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["id", "name", "extension", "itemSize", "date", "parentFolderID"],
        )
        writer.writeheader()
        for item in misc_files:
            writer.writerow(
                {
                    "id": item.get("id"),
                    "name": item.get("name"),
                    "extension": item.get("extension"),
                    "itemSize": item.get("itemSize"),
                    "date": item.get("date"),
                    "parentFolderID": item.get("parentFolderID"),
                }
            )

    print(f"Wrote {len(rows)} textgrid rows to {csv_path}", flush=True)
    print(f"Wrote {len(misc_files)} root-level files to {files_csv}", flush=True)
    if args.max_speakers is not None or args.max_files is not None:
        print(
            "Note: subset run (--max-speakers / --max-files); "
            "full corpus has 202 speaker folders on Box.",
            flush=True,
        )


if __name__ == "__main__":
    main()
